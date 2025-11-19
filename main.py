import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from time import sleep
from datetime import datetime

def parse_habr_python_hub_xlsx():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    articles_data = []
    total_articles = 0

    for page in range(1, 11):
        url = f"https://habr.com/ru/hubs/python/articles/page{page}"
        print(f"Обработка страницы {page}...")

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"Ошибка при запросе к {url}: {e}")
            break

        soup = BeautifulSoup(response.text, 'html.parser')

        articles = soup.select('article.tm-articles-list__item')

        if not articles:
            print(f"На странице {page} статьи не найдены. Останавливаемся.")
            break

        for i, article in enumerate(articles, start=total_articles + 1):
            title_element = article.select_one('h2.tm-title a.tm-title__link')
            if title_element:
                title = title_element.get_text(strip=True)
                link = title_element['href']
                if link.startswith('/'):
                    link = 'https://habr.com' + link
            else:
                title = "Заголовок не найден"
                link = "Ссылка не найдена"

            time_element = article.select_one('time')
            if time_element:
                datetime_str = time_element.get('datetime')
                if datetime_str:
                    try:
                        dt_obj = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
                        formatted_time = dt_obj.strftime('%H:%M, %d.%m.%Y')
                        publication_time = formatted_time
                    except ValueError:
                        publication_time = time_element.get_text(strip=True)
                        print(f"Предупреждение: Не удалось распознать формат даты '{datetime_str}' для статьи '{title}'. Используется текст: '{publication_time}'")
                else:
                    publication_time = time_element.get_text(strip=True)
            else:
                publication_time = "Дата не найдена"

            
            score_element = article.select_one('span.tm-votes-meter__value')
            if score_element:
                score_text = score_element.get_text(strip=True)
                try:
                    score = int(score_text.replace('+', '').replace('-', '').strip())
                    if score_text.startswith('-'):
                        score = -score
                except ValueError:
                    score = 0  
            else:
                score = 0  
            

            articles_data.append([i, title, publication_time, link, score])

        total_articles = len(articles_data)
        print(f"Найдено {len(articles)} статей на странице {page}. Всего: {total_articles}.")

        sleep(1)


    if not articles_data:
        print("Данные для экспорта не найдены.")
        return

    # Создание и форматирование Excel файла
    filename = 'habr_articles_10_pages.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Habr Python Articles'

    headers = ['№', 'Название', 'Дата и время публикации', 'Ссылка', 'Рейтинг статьи']
    ws.append(headers)

    for row_data in articles_data:
        ws.append(row_data)

    # Форматирование заголовков
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center')  

    # Определение стиля границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применение форматирования ко всем ячейкам (кроме заголовков)
    for row in ws.iter_rows(min_row=2): 
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем файл
    try:
        wb.save(filename)
        print(f"Данные успешно экспортированы в файл: {filename}")
        print(f"Всего обработано {total_articles} статей с {page} страниц(ы).")
    except IOError as e:
        print(f"Ошибка при сохранении файла {filename}: {e}")

if __name__ == "__main__":
    parse_habr_python_hub_xlsx()
